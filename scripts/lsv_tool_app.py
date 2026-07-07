#!/usr/bin/env python3
"""
Local drag-and-drop LSV conversion app.

Run:
  python scripts/lsv_tool_app.py

Then open:
  http://127.0.0.1:8765
"""

from __future__ import annotations

import cgi
import csv
import io
import json
import argparse
import math
import time
import webbrowser
import zipfile
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from extract_lsv_vfim import CurveMeta, extract_vfim, sheet_title, write_xlsx


HOST = "127.0.0.1"
PORT = 8765
WORKSPACE_TMP = Path.cwd() / "lsv_vfim_output" / "_tmp"


HTML = r"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>LSV Data Tool</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --text: #1c2430;
      --muted: #647083;
      --line: #d8dde6;
      --accent: #1d7f74;
      --accent-dark: #12665d;
      --soft: #e9f5f2;
      --danger: #b3261e;
      --shadow: 0 18px 50px rgba(30, 38, 50, 0.10);
    }

    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      font-family: "Microsoft YaHei", "Segoe UI", Arial, sans-serif;
      background: var(--bg);
      color: var(--text);
    }

    main {
      width: min(1180px, calc(100vw - 40px));
      margin: 0 auto;
      padding: 34px 0 44px;
    }

    header {
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      gap: 20px;
      margin-bottom: 22px;
    }

    h1 {
      margin: 0 0 8px;
      font-size: 30px;
      line-height: 1.2;
      font-weight: 760;
      letter-spacing: 0;
    }

    .sub {
      margin: 0;
      color: var(--muted);
      font-size: 14px;
      line-height: 1.6;
    }

    .status {
      padding: 10px 14px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
      color: var(--muted);
      font-size: 13px;
      white-space: nowrap;
    }

    .drop {
      display: grid;
      place-items: center;
      min-height: 172px;
      border: 2px dashed #9bb7b2;
      border-radius: 8px;
      background: linear-gradient(180deg, #ffffff 0%, #f0faf7 100%);
      box-shadow: var(--shadow);
      margin-bottom: 18px;
      transition: border-color 140ms ease, transform 140ms ease, background 140ms ease;
    }

    .drop.drag {
      border-color: var(--accent);
      transform: translateY(-1px);
      background: #e9f8f4;
    }

    .drop-inner {
      text-align: center;
      padding: 20px;
    }

    .drop-title {
      font-size: 18px;
      font-weight: 700;
      margin-bottom: 10px;
    }

    .drop-text {
      color: var(--muted);
      font-size: 13px;
      margin-bottom: 18px;
    }

    button, .file-button {
      border: 0;
      border-radius: 8px;
      background: var(--accent);
      color: #fff;
      min-height: 38px;
      padding: 0 16px;
      font-size: 14px;
      font-weight: 650;
      cursor: pointer;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      gap: 8px;
    }

    button:hover, .file-button:hover { background: var(--accent-dark); }
    button.secondary {
      background: #edf1f5;
      color: var(--text);
      border: 1px solid var(--line);
    }
    button.secondary:hover { background: #e1e7ed; }
    button.danger {
      background: #fff2f1;
      color: var(--danger);
      border: 1px solid #f1c5c1;
    }
    button.danger:hover { background: #ffe5e2; }
    button:disabled {
      opacity: .55;
      cursor: not-allowed;
    }

    input[type="file"] { display: none; }

    .toolbar {
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 16px;
      align-items: end;
      margin: 18px 0;
    }

    .defaults {
      display: flex;
      flex-wrap: wrap;
      gap: 12px;
      align-items: end;
    }

    label {
      display: grid;
      gap: 6px;
      color: var(--muted);
      font-size: 12px;
      font-weight: 650;
    }

    select {
      width: 132px;
      height: 38px;
      border-radius: 8px;
      border: 1px solid var(--line);
      background: #fff;
      padding: 0 10px;
      color: var(--text);
      font-size: 14px;
      outline: none;
    }

    input {
      width: 116px;
      height: 38px;
      border-radius: 8px;
      border: 1px solid var(--line);
      background: #fff;
      padding: 0 10px;
      color: var(--text);
      font-size: 14px;
      outline: none;
    }

    input:focus {
      border-color: var(--accent);
      box-shadow: 0 0 0 3px rgba(29, 127, 116, .14);
    }

    .table-wrap {
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
      box-shadow: var(--shadow);
    }

    table {
      width: 100%;
      border-collapse: collapse;
      min-width: 780px;
    }

    th, td {
      border-bottom: 1px solid var(--line);
      padding: 10px 12px;
      text-align: left;
      font-size: 13px;
      vertical-align: middle;
    }

    th {
      position: sticky;
      top: 0;
      background: #f9fbfc;
      color: #3f4b5a;
      font-size: 12px;
      letter-spacing: 0;
      z-index: 1;
    }

    tr:last-child td { border-bottom: 0; }
    td.name {
      max-width: 360px;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
      font-weight: 650;
    }

    td.size { color: var(--muted); width: 92px; }
    td.actions { width: 82px; text-align: right; }
    td input { width: 100%; min-width: 92px; }
    td.sheet input { min-width: 160px; }

    .empty {
      text-align: center;
      padding: 36px 14px;
      color: var(--muted);
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
    }

    .footer {
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 14px;
      margin-top: 18px;
    }

    .note {
      color: var(--muted);
      font-size: 13px;
      line-height: 1.5;
    }

    .error {
      color: var(--danger);
      font-size: 13px;
      margin-top: 10px;
      min-height: 20px;
    }

    @media (max-width: 760px) {
      main { width: min(100vw - 24px, 1180px); padding-top: 22px; }
      header, .toolbar, .footer { display: grid; }
      .status { white-space: normal; }
      .defaults { display: grid; grid-template-columns: 1fr 1fr; }
      select {
      width: 132px;
      height: 38px;
      border-radius: 8px;
      border: 1px solid var(--line);
      background: #fff;
      padding: 0 10px;
      color: var(--text);
      font-size: 14px;
      outline: none;
    }

    input { width: 100%; }
      h1 { font-size: 24px; }
    }
  </style>
</head>
<body>
  <main>
    <header>
      <div>
        <h1>LSV Data Tool</h1>
        <p class="sub">Drop raw .DTA files, enter pH and electrode area, then export processed LSV, Tafel, target overpotential tables, and plots.</p>
      </div>
      <div class="status" id="status">Running locally. Files stay on this computer.</div>
    </header>

    <section class="drop" id="drop">
      <div class="drop-inner">
        <div class="drop-title">Drop raw LSV files here</div>
        <div class="drop-text">Multiple files are supported. You can also click the button to choose files.</div>
        <label class="file-button" for="fileInput">Choose files</label>
        <input id="fileInput" type="file" multiple />
      </div>
    </section>

    <section class="toolbar">
      <div class="defaults">
        <label>Tafel min j
          <input id="tafelMinJ" type="number" step="0.1" value="10" />
        </label>
        <label>Tafel max j
          <input id="tafelMaxJ" type="number" step="0.1" value="100" />
        </label>
        <label>Default pH
          <input id="defaultPh" type="number" step="0.001" value="13.2" />
        </label>
        <label>Default area cm2
          <input id="defaultArea" type="number" step="0.0001" value="0.25" />
        </label>
        <label>Reference
          <select id="referenceElectrode">
            <option value="Ag/AgCl" selected>Ag/AgCl</option>
            <option value="Hg/HgO">Hg/HgO</option>
            <option value="Custom">Custom</option>
          </select>
        </label>
        <label>Ref to SHE V
          <input id="refToShe" type="number" step="0.001" value="0.197" />
        </label>
        <label>iR correction
          <select id="applyIR">
            <option value="false" selected>False</option>
            <option value="true">True</option>
          </select>
        </label>
        <label>Rs ohm
          <input id="rsOhm" type="number" step="0.0001" value="0" />
        </label>
        <label>iR fraction
          <input id="irFraction" type="number" step="0.01" value="1.0" />
        </label>
        <label>Current sign
          <select id="currentSignMode">
            <option value="auto" selected>auto</option>
            <option value="positive">positive</option>
            <option value="negative">negative</option>
          </select>
        </label>
        <label>Target j list
          <input id="targetJs" type="text" value="100,200,300" />
        </label>
        <button class="secondary" id="applyDefaults" type="button">Fill blanks</button>
      </div>
      <button class="danger" id="clearFiles" type="button">Clear</button>
    </section>

    <div id="fileArea" class="empty">No files yet. Drop or choose .DTA files, then enter parameters for each curve.</div>

    <div class="footer">
      <div class="note">Exports include full Excel data, Origin-ready CE tables, LSV/Tafel plots, and eta targets at selected current densities.</div>
      <button id="exportBtn" type="button" disabled>Export results</button>
    </div>
    <div class="error" id="error"></div>
  </main>

  <script>
    const state = new Map();
    const drop = document.getElementById('drop');
    const input = document.getElementById('fileInput');
    const fileArea = document.getElementById('fileArea');
    const exportBtn = document.getElementById('exportBtn');
    const errorBox = document.getElementById('error');
    const statusBox = document.getElementById('status');

    function safeSheetName(name) {
      return name.replace(/\.[^.]+$/g, '').replace(/[\[\]:*?\/\\]/g, '_').slice(0, 31) || 'LSV';
    }

    function formatSize(size) {
      if (size < 1024) return `${size} B`;
      if (size < 1024 * 1024) return `${(size / 1024).toFixed(1)} KB`;
      return `${(size / 1024 / 1024).toFixed(2)} MB`;
    }

    function addFiles(files) {
      for (const file of files) {
        state.set(file.name, {
          file,
          ph: document.getElementById('defaultPh').value || '',
          area: document.getElementById('defaultArea').value || '',
          sheetName: safeSheetName(file.name)
        });
      }
      render();
    }

    function render() {
      errorBox.textContent = '';
      exportBtn.disabled = state.size === 0;
      statusBox.textContent = state.size ? `Added ${state.size} file(s)` : 'Running locally. Files stay on this computer.';
      if (!state.size) {
        fileArea.className = 'empty';
        fileArea.textContent = 'No files yet. Drop or choose .DTA files, then enter parameters for each curve.';
        return;
      }

      fileArea.className = 'table-wrap';
      const rows = [...state.values()].map((item) => `
        <tr>
          <td class="name" title="${item.file.name}">${item.file.name}</td>
          <td class="size">${formatSize(item.file.size)}</td>
          <td><input data-field="ph" data-name="${item.file.name}" type="number" step="0.001" value="${item.ph}" placeholder="pH"></td>
          <td><input data-field="area" data-name="${item.file.name}" type="number" step="0.0001" value="${item.area}" placeholder="cm2"></td>
          <td class="sheet"><input data-field="sheetName" data-name="${item.file.name}" value="${item.sheetName}" placeholder="sheet name"></td>
          <td class="actions"><button class="danger" data-remove="${item.file.name}" type="button">Remove</button></td>
        </tr>
      `).join('');

      fileArea.innerHTML = `
        <table>
          <thead>
            <tr>
              <th>File</th>
              <th>Size</th>
              <th>pH</th>
              <th>Area cm2</th>
              <th>Sheet name</th>
              <th></th>
            </tr>
          </thead>
          <tbody>${rows}</tbody>
        </table>
      `;

      fileArea.querySelectorAll('input').forEach((el) => {
        el.addEventListener('input', () => {
          const item = state.get(el.dataset.name);
          item[el.dataset.field] = el.value;
        });
      });
      fileArea.querySelectorAll('button[data-remove]').forEach((btn) => {
        btn.addEventListener('click', () => {
          state.delete(btn.dataset.remove);
          render();
        });
      });
    }

    ['dragenter', 'dragover'].forEach((eventName) => {
      drop.addEventListener(eventName, (event) => {
        event.preventDefault();
        drop.classList.add('drag');
      });
    });
    ['dragleave', 'drop'].forEach((eventName) => {
      drop.addEventListener(eventName, (event) => {
        event.preventDefault();
        drop.classList.remove('drag');
      });
    });
    drop.addEventListener('drop', (event) => addFiles(event.dataTransfer.files));
    input.addEventListener('change', () => addFiles(input.files));

    document.getElementById('applyDefaults').addEventListener('click', () => {
      const ph = document.getElementById('defaultPh').value;
      const area = document.getElementById('defaultArea').value;
      for (const item of state.values()) {
        if (!item.ph && ph) item.ph = ph;
        if (!item.area && area) item.area = area;
      }
      render();
    });

    document.getElementById('referenceElectrode').addEventListener('change', () => {
      const ref = document.getElementById('referenceElectrode').value;
      const refToShe = document.getElementById('refToShe');
      if (ref === 'Ag/AgCl') refToShe.value = '0.197';
      if (ref === 'Hg/HgO') refToShe.value = '0.098';
    });
    document.getElementById('clearFiles').addEventListener('click', () => {
      state.clear();
      input.value = '';
      render();
    });

    exportBtn.addEventListener('click', async () => {
      errorBox.textContent = '';
      const items = [...state.values()];
      for (const item of items) {
        if (!item.ph || !item.area) {
          errorBox.textContent = 'Please enter pH and area for every file.';
          return;
        }
      }

      const form = new FormData();
      const tafelMinJ = Number(document.getElementById('tafelMinJ').value || 10);
      const tafelMaxJ = Number(document.getElementById('tafelMaxJ').value || 100);
      const refToShe = Number(document.getElementById('refToShe').value);
      const rsOhm = Number(document.getElementById('rsOhm').value || 0);
      const irFraction = Number(document.getElementById('irFraction').value || 1);
      const targetJs = document.getElementById('targetJs').value || '100,200,300';
      if (!(tafelMinJ > 0) || !(tafelMaxJ > tafelMinJ)) {
        errorBox.textContent = 'Tafel range must satisfy: 0 < min j < max j.';
        return;
      }
      if (!Number.isFinite(refToShe) || !Number.isFinite(rsOhm) || !Number.isFinite(irFraction)) {
        errorBox.textContent = 'Reference/iR parameters must be numeric.';
        return;
      }
      const manifest = items.map((item) => ({
        name: item.file.name,
        pH: Number(item.ph),
        area_cm2: Number(item.area),
        sheet_name: item.sheetName || safeSheetName(item.file.name)
      }));
      form.append('manifest', JSON.stringify(manifest));
      form.append('tafel_min_j', String(tafelMinJ));
      form.append('tafel_max_j', String(tafelMaxJ));
      form.append('reference_electrode', document.getElementById('referenceElectrode').value);
      form.append('ref_to_she_v', String(refToShe));
      form.append('apply_ir_correction', document.getElementById('applyIR').value);
      form.append('rs_ohm', String(rsOhm));
      form.append('ir_compensation_fraction', String(irFraction));
      form.append('current_sign_mode', document.getElementById('currentSignMode').value);
      form.append('target_current_densities', targetJs);
      for (const item of items) form.append('files', item.file, item.file.name);

      exportBtn.disabled = true;
      exportBtn.textContent = 'Generating...';
      try {
        const response = await fetch('/api/export', { method: 'POST', body: form });
        if (!response.ok) {
          const text = await response.text();
          throw new Error(text || `HTTP ${response.status}`);
        }
        const blob = await response.blob();
        const url = URL.createObjectURL(blob);
        const link = document.createElement('a');
        link.href = url;
        link.download = `LSV_export_${new Date().toISOString().slice(0, 19).replace(/[:T]/g, '')}.zip`;
        document.body.appendChild(link);
        link.click();
        link.remove();
        URL.revokeObjectURL(url);
        statusBox.textContent = 'Export complete';
      } catch (error) {
        errorBox.textContent = error.message || String(error);
      } finally {
        exportBtn.disabled = state.size === 0;
        exportBtn.textContent = 'Export results';
      }
    });
  </script>
</body>
</html>
"""


def build_origin_tables(
    curves: list[tuple[Path, list[Any], CurveMeta]],
    resistance_ohm: float = 3.8504,
    rhe_offset_v: float = 0.197,
    ph_slope_v: float = 0.0591,
) -> tuple[bytes, bytes]:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    used_titles: set[str] = set()
    rows_by_curve: list[tuple[str, list[tuple[float, float]]]] = []
    max_rows = 0

    for source, points, meta in curves:
        title = sheet_title(meta.sheet_name or source.name, used_titles)
        rhe_shift = rhe_offset_v + ph_slope_v * meta.ph
        rows = [
            (
                point.vf_v + rhe_shift,
                point.im_a * 1000 / meta.area_cm2,
            )
            for point in points
        ]
        rows_by_curve.append((title, rows))
        max_rows = max(max_rows, len(rows))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Origin_CE"

    csv_buffer = io.StringIO(newline="")
    csv_writer = csv.writer(csv_buffer)

    header: list[str] = []
    for title, _rows in rows_by_curve:
        header.extend([f"{title}_v_V_vs_RHE", f"{title}_current_density_mA_cm2"])
    sheet.append(header)
    csv_writer.writerow(header)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row_index in range(max_rows):
        row: list[float | str] = []
        for _title, rows in rows_by_curve:
            if row_index < len(rows):
                row.extend(rows[row_index])
            else:
                row.extend(["", ""])
        sheet.append(row)
        csv_writer.writerow(row)

    for col_index in range(1, len(header) + 1):
        sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = 24

    xlsx_buffer = io.BytesIO()
    workbook.save(xlsx_buffer)
    return xlsx_buffer.getvalue(), csv_buffer.getvalue().encode("utf-8-sig")


def nice_number(value: float, round_value: bool) -> float:
    if value <= 0:
        return 1.0
    exponent = math.floor(math.log10(value))
    fraction = value / 10**exponent
    if round_value:
        if fraction < 1.5:
            nice_fraction = 1
        elif fraction < 3:
            nice_fraction = 2
        elif fraction < 7:
            nice_fraction = 5
        else:
            nice_fraction = 10
    else:
        if fraction <= 1:
            nice_fraction = 1
        elif fraction <= 2:
            nice_fraction = 2
        elif fraction <= 5:
            nice_fraction = 5
        else:
            nice_fraction = 10
    return nice_fraction * 10**exponent


def nice_ticks(min_value: float, max_value: float, target_count: int = 5) -> list[float]:
    if math.isclose(min_value, max_value):
        span = abs(max_value) or 1
        min_value -= span * 0.5
        max_value += span * 0.5
    span = nice_number(max_value - min_value, False)
    step = nice_number(span / max(target_count - 1, 1), True)
    tick_min = math.floor(min_value / step) * step
    tick_max = math.ceil(max_value / step) * step
    ticks = []
    value = tick_min
    guard = 0
    while value <= tick_max + step * 0.5 and guard < 100:
        ticks.append(round(value, 10))
        value += step
        guard += 1
    return ticks


def format_tick(value: float) -> str:
    if abs(value) >= 100:
        return f"{value:.0f}"
    if abs(value) >= 10:
        return f"{value:.1f}".rstrip("0").rstrip(".")
    return f"{value:.2f}".rstrip("0").rstrip(".")


def find_font(size: int):
    from PIL import ImageFont

    candidates = [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/simsun.ttc",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, size=size)
    return ImageFont.load_default()


def draw_text_centered(draw: Any, position: tuple[float, float], text: str, font: Any, fill: str) -> None:
    bbox = draw.textbbox((0, 0), text, font=font)
    width = bbox[2] - bbox[0]
    height = bbox[3] - bbox[1]
    draw.text((position[0] - width / 2, position[1] - height / 2), text, font=font, fill=fill)


def draw_rotated_label(image: Any, text: str, center: tuple[int, int], font: Any, fill: str) -> None:
    from PIL import Image, ImageDraw

    bbox = font.getbbox(text)
    width = bbox[2] - bbox[0] + 16
    height = bbox[3] - bbox[1] + 16
    label = Image.new("RGBA", (width, height), (0, 0, 0, 0))
    label_draw = ImageDraw.Draw(label)
    label_draw.text((8 - bbox[0], 8 - bbox[1]), text, font=font, fill=fill)
    rotated = label.rotate(90, expand=True)
    image.alpha_composite(rotated, (int(center[0] - rotated.width / 2), int(center[1] - rotated.height / 2)))


def build_lsv_plot_png(
    curves: list[tuple[Path, list[Any], CurveMeta]],
    resistance_ohm: float = 3.8504,
    rhe_offset_v: float = 0.197,
    ph_slope_v: float = 0.0591,
) -> bytes:
    from PIL import Image, ImageDraw

    palette = ["#6a6a6a", "#ff4d4d", "#2f80ed", "#43bf79", "#b46ae5", "#e2b500", "#00a6a6", "#f28e2b"]
    plot_curves: list[tuple[str, str, list[tuple[float, float]]]] = []
    all_x: list[float] = []
    all_y: list[float] = []
    used_titles: set[str] = set()

    for index, (source, points, meta) in enumerate(curves):
        title = sheet_title(meta.sheet_name or source.name, used_titles)
        rhe_shift = rhe_offset_v + ph_slope_v * meta.ph
        rows = [
            (
                point.vf_v + rhe_shift,
                point.im_a * 1000 / meta.area_cm2,
            )
            for point in points
        ]
        rows = [(x, y) for x, y in rows if math.isfinite(x) and math.isfinite(y)]
        if not rows:
            continue
        plot_curves.append((title, palette[index % len(palette)], rows))
        all_x.extend(x for x, _y in rows)
        all_y.extend(y for _x, y in rows)

    if not plot_curves:
        raise ValueError("No drawable curve data.")

    width, height = 1800, 1400
    left, right, top, bottom = 290, 120, 150, 190
    plot_left, plot_top = left, top
    plot_right, plot_bottom = width - right, height - bottom
    plot_width = plot_right - plot_left
    plot_height = plot_bottom - plot_top

    # Match the user's reference plotting window.
    x_min, x_max = 1.0, 1.8
    y_min, y_max = 0.0, 250.0
    x_ticks = [1.0, 1.2, 1.4, 1.6, 1.8]
    y_ticks = [0.0, 100.0, 200.0]

    def x_to_px(x_value: float) -> float:
        return plot_left + (x_value - x_min) / (x_max - x_min) * plot_width

    def y_to_px(y_value: float) -> float:
        return plot_bottom - (y_value - y_min) / (y_max - y_min) * plot_height

    image = Image.new("RGBA", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)
    font_tick = find_font(54)
    font_label = find_font(68)
    font_legend = find_font(40)
    font_small = find_font(34)

    axis_color = "#000000"
    tick_color = "#000000"
    label_color = "#000000"
    draw.line((plot_left, plot_bottom, plot_right, plot_bottom), fill=axis_color, width=6)
    draw.line((plot_left, plot_top, plot_left, plot_bottom), fill=axis_color, width=6)

    for tick in x_ticks:
        x = x_to_px(tick)
        draw.line((x, plot_bottom, x, plot_bottom + 12), fill=tick_color, width=5)
        draw_text_centered(draw, (x, plot_bottom + 62), format_tick(tick), font_tick, label_color)

    for tick in y_ticks:
        y = y_to_px(tick)
        draw.line((plot_left - 12, y, plot_left, y), fill=tick_color, width=5)
        bbox = draw.textbbox((0, 0), format_tick(tick), font=font_tick)
        draw.text((plot_left - 32 - (bbox[2] - bbox[0]), y - (bbox[3] - bbox[1]) / 2), format_tick(tick), font=font_tick, fill=label_color)

    draw_text_centered(draw, ((plot_left + plot_right) / 2, height - 68), "Potential / V vs RHE", font_label, label_color)
    draw_rotated_label(image, "j/(mA cm-2)", (76, (plot_top + plot_bottom) // 2), font_label, label_color)

    for title, color, rows in plot_curves:
        visible_rows = [(x, y) for x, y in rows if x_min <= x <= x_max and y_min <= y <= y_max]
        points = [(x_to_px(x), y_to_px(y)) for x, y in visible_rows]
        if len(points) >= 2:
            draw.line(points, fill=color, width=12, joint="curve")

    legend_x, legend_y = 340, 210
    line_len = 130
    row_gap = 64
    for index, (title, color, _rows) in enumerate(plot_curves):
        y = legend_y + index * row_gap
        draw.line((legend_x, y, legend_x + line_len, y), fill=color, width=13)
        draw.text((legend_x + line_len + 20, y - 25), title, font=font_legend, fill=label_color)

    note_lines = [
        "Working electrode: auto exported curve",
        "Counter electrode: carbon rod",
        "Reference electrode: see inputs",
    ]
    note_y = legend_y + len(plot_curves) * row_gap + 30
    for line in note_lines:
        draw.text((legend_x, note_y), line, font=font_small, fill=label_color)
        note_y += 50

    output = io.BytesIO()
    image.convert("RGB").save(output, format="PNG", optimize=True)
    return output.getvalue()


def linear_regression(points: list[tuple[float, float]]) -> tuple[float, float, float]:
    if len(points) < 2:
        raise ValueError("At least two points are required for linear regression.")
    n = len(points)
    sum_x = sum(x for x, _y in points)
    sum_y = sum(y for _x, y in points)
    mean_x = sum_x / n
    mean_y = sum_y / n
    ss_xx = sum((x - mean_x) ** 2 for x, _y in points)
    if math.isclose(ss_xx, 0.0):
        raise ValueError("X values are identical; cannot fit Tafel slope.")
    ss_xy = sum((x - mean_x) * (y - mean_y) for x, y in points)
    slope = ss_xy / ss_xx
    intercept = mean_y - slope * mean_x
    ss_tot = sum((y - mean_y) ** 2 for _x, y in points)
    ss_res = sum((y - (slope * x + intercept)) ** 2 for x, y in points)
    r_squared = 1.0 if math.isclose(ss_tot, 0.0) else 1.0 - ss_res / ss_tot
    return slope, intercept, r_squared


def compute_tafel_fits(
    curves: list[tuple[Path, list[Any], CurveMeta]],
    min_j: float,
    max_j: float,
    rhe_offset_v: float = 0.197,
    ph_slope_v: float = 0.0591,
) -> list[dict[str, Any]]:
    if not (min_j > 0 and max_j > min_j):
        raise ValueError("Tafel fit range must satisfy: 0 < min j < max j.")

    used_titles: set[str] = set()
    results: list[dict[str, Any]] = []
    for source, points, meta in curves:
        title = sheet_title(meta.sheet_name or source.name, used_titles)
        rhe_shift = rhe_offset_v + ph_slope_v * meta.ph
        all_rows: list[tuple[float, float, float, float]] = []
        fit_points: list[tuple[float, float]] = []

        for point in points:
            e_rhe = point.vf_v + rhe_shift
            current_density = point.im_a * 1000 / meta.area_cm2
            if current_density <= 0:
                continue
            overpotential_mv = (e_rhe - 1.23) * 1000
            log_j = math.log10(current_density)
            all_rows.append((e_rhe, current_density, log_j, overpotential_mv))
            if min_j <= current_density <= max_j:
                fit_points.append((log_j, overpotential_mv))

        slope = intercept = r_squared = None
        status = "OK"
        if len(fit_points) >= 2:
            slope, intercept, r_squared = linear_regression(fit_points)
        else:
            status = "Not enough points in selected j range"

        results.append(
            {
                "title": title,
                "source_file": str(source),
                "ph": meta.ph,
                "area_cm2": meta.area_cm2,
                "min_j": min_j,
                "max_j": max_j,
                "all_rows": all_rows,
                "fit_points": fit_points,
                "slope_mv_dec": slope,
                "intercept_mv": intercept,
                "r_squared": r_squared,
                "fit_count": len(fit_points),
                "status": status,
            }
        )
    return results


def build_tafel_workbook(fits: list[dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Tafel_summary"
    summary.append(
        [
            "curve",
            "source_file",
            "pH",
            "area_cm2",
            "fit_min_j_mA_cm2",
            "fit_max_j_mA_cm2",
            "points_used",
            "slope_mV_dec",
            "intercept_mV",
            "R_squared",
            "status",
        ]
    )
    for cell in summary[1]:
        cell.font = Font(bold=True)

    used_sheets = {summary.title}
    for fit in fits:
        summary.append(
            [
                fit["title"],
                fit["source_file"],
                fit["ph"],
                fit["area_cm2"],
                fit["min_j"],
                fit["max_j"],
                fit["fit_count"],
                fit["slope_mv_dec"],
                fit["intercept_mv"],
                fit["r_squared"],
                fit["status"],
            ]
        )

        sheet = workbook.create_sheet(sheet_title(f"Tafel_{fit['title']}", used_sheets))
        sheet.append(["E_RHE_V", "j_mA_cm2", "log10_j", "eta_mV", "used_for_fit"])
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        selected = set(fit["fit_points"])
        for e_rhe, current_density, log_j, overpotential_mv in fit["all_rows"]:
            sheet.append(
                [
                    e_rhe,
                    current_density,
                    log_j,
                    overpotential_mv,
                    "Y" if (log_j, overpotential_mv) in selected else "",
                ]
            )
        for column, width in {"A": 14, "B": 16, "C": 14, "D": 14, "E": 14}.items():
            sheet.column_dimensions[column].width = width
        sheet.freeze_panes = "A2"

    for column, width in {
        "A": 22,
        "B": 64,
        "C": 10,
        "D": 12,
        "E": 16,
        "F": 16,
        "G": 13,
        "H": 16,
        "I": 16,
        "J": 12,
        "K": 28,
    }.items():
        summary.column_dimensions[column].width = width
    summary.freeze_panes = "A2"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_tafel_plot_png(fits: list[dict[str, Any]]) -> bytes:
    from PIL import Image, ImageDraw

    plot_series = [fit for fit in fits if fit["fit_points"]]
    if not plot_series:
        raise ValueError("No Tafel points available for plotting.")

    palette = ["#6a6a6a", "#ff4d4d", "#2f80ed", "#43bf79", "#b46ae5", "#e2b500", "#00a6a6", "#f28e2b"]
    all_x = [x for fit in plot_series for x, _y in fit["fit_points"]]
    all_y = [y for fit in plot_series for _x, y in fit["fit_points"]]
    for fit in plot_series:
        if fit["slope_mv_dec"] is not None:
            all_y.extend(
                [
                    fit["slope_mv_dec"] * min(all_x) + fit["intercept_mv"],
                    fit["slope_mv_dec"] * max(all_x) + fit["intercept_mv"],
                ]
            )

    width, height = 1800, 1400
    left, right, top, bottom = 300, 120, 150, 190
    plot_left, plot_top = left, top
    plot_right, plot_bottom = width - right, height - bottom
    plot_width = plot_right - plot_left
    plot_height = plot_bottom - plot_top

    x_ticks = nice_ticks(min(all_x), max(all_x), 5)
    y_ticks = nice_ticks(min(all_y), max(all_y), 5)
    x_min, x_max = min(x_ticks), max(x_ticks)
    y_min, y_max = min(y_ticks), max(y_ticks)

    def x_to_px(x_value: float) -> float:
        return plot_left + (x_value - x_min) / (x_max - x_min) * plot_width

    def y_to_px(y_value: float) -> float:
        return plot_bottom - (y_value - y_min) / (y_max - y_min) * plot_height

    image = Image.new("RGBA", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)
    font_tick = find_font(50)
    font_label = find_font(64)
    font_legend = find_font(38)
    axis_color = "#000000"
    label_color = "#000000"

    draw.line((plot_left, plot_bottom, plot_right, plot_bottom), fill=axis_color, width=6)
    draw.line((plot_left, plot_top, plot_left, plot_bottom), fill=axis_color, width=6)

    for tick in x_ticks:
        x = x_to_px(tick)
        draw.line((x, plot_bottom, x, plot_bottom + 12), fill=axis_color, width=5)
        draw_text_centered(draw, (x, plot_bottom + 62), format_tick(tick), font_tick, label_color)
    for tick in y_ticks:
        y = y_to_px(tick)
        draw.line((plot_left - 12, y, plot_left, y), fill=axis_color, width=5)
        label = format_tick(tick)
        bbox = draw.textbbox((0, 0), label, font=font_tick)
        draw.text((plot_left - 32 - (bbox[2] - bbox[0]), y - (bbox[3] - bbox[1]) / 2), label, font=font_tick, fill=label_color)

    draw_text_centered(draw, ((plot_left + plot_right) / 2, height - 68), "log10 j/(mA cm-2)", font_label, label_color)
    draw_rotated_label(image, "eta / mV", (78, (plot_top + plot_bottom) // 2), font_label, label_color)

    legend_x, legend_y = 340, 210
    for index, fit in enumerate(plot_series):
        color = palette[index % len(palette)]
        points = [(x_to_px(x), y_to_px(y)) for x, y in fit["fit_points"]]
        for x, y in points:
            draw.ellipse((x - 7, y - 7, x + 7, y + 7), fill=color)
        if fit["slope_mv_dec"] is not None:
            x1 = min(x for x, _y in fit["fit_points"])
            x2 = max(x for x, _y in fit["fit_points"])
            y1 = fit["slope_mv_dec"] * x1 + fit["intercept_mv"]
            y2 = fit["slope_mv_dec"] * x2 + fit["intercept_mv"]
            draw.line((x_to_px(x1), y_to_px(y1), x_to_px(x2), y_to_px(y2)), fill=color, width=9)

        legend_yi = legend_y + index * 62
        label = fit["title"]
        if fit["slope_mv_dec"] is not None:
            label = f"{label}: {fit['slope_mv_dec']:.1f} mV/dec, R2={fit['r_squared']:.4f}"
        draw.line((legend_x, legend_yi, legend_x + 120, legend_yi), fill=color, width=10)
        draw.text((legend_x + 145, legend_yi - 24), label, font=font_legend, fill=label_color)

    output = io.BytesIO()
    image.convert("RGB").save(output, format="PNG", optimize=True)
    return output.getvalue()


def finite_or_none(value: float | None) -> float | None:
    if value is None:
        return None
    if not isinstance(value, (int, float)) or not math.isfinite(value):
        return None
    return float(value)


def infer_current_multiplier(points: list[Any], mode: str) -> tuple[float, str]:
    mode = (mode or "auto").lower()
    if mode == "positive":
        return 1.0, "OK"
    if mode == "negative":
        return -1.0, "OK"
    if mode != "auto":
        return 1.0, f"warning: unknown current_sign_mode {mode}; assumed positive"

    currents = [point.im_a for point in points if math.isfinite(point.im_a)]
    if not currents:
        return 1.0, "warning: no finite current values; assumed positive"
    max_pos = max([current for current in currents if current > 0], default=0.0)
    max_neg_abs = abs(min([current for current in currents if current < 0], default=0.0))
    if max_pos > max_neg_abs * 1.1:
        return 1.0, "OK"
    if max_neg_abs > max_pos * 1.1:
        return -1.0, "OK"
    return 1.0, "warning: current sign ambiguous; assumed positive"


def interpolate_linear(target_x: float, xs: list[float], ys: list[float]) -> float | None:
    if not xs or len(xs) != len(ys):
        return None
    if target_x < xs[0] or target_x > xs[-1]:
        return None
    for index, x_value in enumerate(xs):
        if math.isclose(target_x, x_value, rel_tol=1e-9, abs_tol=1e-9):
            return ys[index]
    for index in range(1, len(xs)):
        x0, x1 = xs[index - 1], xs[index]
        if x0 <= target_x <= x1 and not math.isclose(x0, x1):
            y0, y1 = ys[index - 1], ys[index]
            return y0 + (target_x - x0) * (y1 - y0) / (x1 - x0)
    return None


def extract_overpotentials_at_targets(
    points: list[Any],
    sample_name: str,
    meta: CurveMeta,
    target_current_densities: list[float],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    multiplier, sign_status = infer_current_multiplier(points, meta.current_sign_mode)
    processed_rows: list[dict[str, Any]] = []
    positive_rows: list[dict[str, Any]] = []

    for point in points:
        e_ref_v = finite_or_none(point.vf_v)
        current_a = finite_or_none(point.im_a)
        if e_ref_v is None or current_a is None or meta.area_cm2 <= 0:
            continue
        j_mA_cm2 = multiplier * current_a * 1000 / meta.area_cm2
        e_rhe_v = e_ref_v + meta.ref_to_she_v + 0.0591 * meta.ph
        if meta.apply_ir_correction:
            e_rhe_ircorr_v = e_rhe_v - current_a * meta.rs_ohm * meta.ir_compensation_fraction
        else:
            e_rhe_ircorr_v = e_rhe_v
        eta_v = e_rhe_ircorr_v - 1.23
        eta_mV = eta_v * 1000
        row = {
            "Sample": sample_name,
            "Pt": getattr(point, "point", None),
            "Time_s": getattr(point, "time_s", None),
            "E_ref_V": e_ref_v,
            "I_A": current_a,
            "Current_sign_multiplier": multiplier,
            "j_mA_cm2": j_mA_cm2,
            "E_RHE_V": e_rhe_v,
            "E_RHE_iRcorr_V": e_rhe_ircorr_v,
            "Eta_V": eta_v,
            "Eta_mV": eta_mV,
            "OER_positive_branch": j_mA_cm2 > 0,
            "Status": sign_status,
        }
        processed_rows.append(row)
        if j_mA_cm2 > 0 and math.isfinite(j_mA_cm2) and math.isfinite(eta_mV):
            positive_rows.append(row)

    # Sort by j and merge repeated j values by rounded bins to keep interpolation stable.
    grouped: dict[float, list[dict[str, Any]]] = {}
    for row in sorted(positive_rows, key=lambda item: item["j_mA_cm2"]):
        grouped.setdefault(round(row["j_mA_cm2"], 9), []).append(row)

    unique_rows: list[dict[str, Any]] = []
    for _j_key, rows in grouped.items():
        count = len(rows)
        unique_rows.append(
            {
                "j_mA_cm2": sum(row["j_mA_cm2"] for row in rows) / count,
                "E_RHE_iRcorr_V": sum(row["E_RHE_iRcorr_V"] for row in rows) / count,
                "Eta_mV": sum(row["Eta_mV"] for row in rows) / count,
            }
        )
    unique_rows.sort(key=lambda item: item["j_mA_cm2"])
    j_sorted = [row["j_mA_cm2"] for row in unique_rows]
    eta_sorted = [row["Eta_mV"] for row in unique_rows]
    e_sorted = [row["E_RHE_iRcorr_V"] for row in unique_rows]

    target_rows: list[dict[str, Any]] = []
    summary: dict[str, Any] = {"Sample": sample_name, "Overpotential_note": sign_status if sign_status != "OK" else "OK"}
    for target_j in target_current_densities:
        eta_mV = interpolate_linear(target_j, j_sorted, eta_sorted)
        e_rhe_ircorr = interpolate_linear(target_j, j_sorted, e_sorted)
        status = "OK"
        if eta_mV is None or e_rhe_ircorr is None:
            status = "out of range"
            if summary["Overpotential_note"] == "OK":
                summary["Overpotential_note"] = "out of range"
            elif "out of range" not in summary["Overpotential_note"]:
                summary["Overpotential_note"] += "; out of range"
        target_label = int(target_j) if float(target_j).is_integer() else target_j
        target_rows.append(
            {
                "Sample": sample_name,
                "Target_j_mA_cm2": target_j,
                "Current_A_equivalent": target_j * meta.area_cm2 / 1000,
                "E_RHE_iRcorr_V": e_rhe_ircorr,
                "Eta_mV": eta_mV,
                "Eta_V": None if eta_mV is None else eta_mV / 1000,
                "pH": meta.ph,
                "Area_cm2": meta.area_cm2,
                "Reference": meta.reference_electrode,
                "Ref_to_SHE_V": meta.ref_to_she_v,
                "iR_corrected": meta.apply_ir_correction,
                "Rs_ohm": meta.rs_ohm,
                "Status": status,
            }
        )
        summary[f"Eta_{target_label}_mV"] = eta_mV
        summary[f"E_RHE_at_{target_label}_mAcm2_V"] = e_rhe_ircorr

    return processed_rows, target_rows, summary


def build_overpotential_plot_png(
    processed_by_sample: dict[str, list[dict[str, Any]]],
    target_rows: list[dict[str, Any]],
    target_current_densities: list[float],
) -> bytes:
    from PIL import Image, ImageDraw

    palette = ["#6a6a6a", "#ff4d4d", "#2f80ed", "#43bf79", "#b46ae5", "#e2b500", "#00a6a6", "#f28e2b"]
    series: list[tuple[str, str, list[tuple[float, float]]]] = []
    all_x: list[float] = []
    all_y: list[float] = []
    for index, (sample, rows) in enumerate(processed_by_sample.items()):
        points_xy = [
            (row["j_mA_cm2"], row["Eta_mV"])
            for row in rows
            if row.get("OER_positive_branch") and math.isfinite(row["j_mA_cm2"]) and math.isfinite(row["Eta_mV"])
        ]
        points_xy.sort(key=lambda item: item[0])
        if points_xy:
            series.append((sample, palette[index % len(palette)], points_xy))
            all_x.extend(x for x, _y in points_xy)
            all_y.extend(y for _x, y in points_xy)

    if not series:
        raise ValueError("No positive OER branch data available for overpotential plot.")

    x_max = max(max(all_x), max(target_current_densities)) * 1.08
    y_min = min(0.0, min(all_y))
    y_max = max(all_y) * 1.08
    x_ticks = nice_ticks(0.0, x_max, 6)
    y_ticks = nice_ticks(y_min, y_max, 6)
    x_min, x_max = min(x_ticks), max(x_ticks)
    y_min, y_max = min(y_ticks), max(y_ticks)

    width, height = 1800, 1400
    left, right, top, bottom = 300, 120, 150, 190
    plot_left, plot_top = left, top
    plot_right, plot_bottom = width - right, height - bottom
    plot_width = plot_right - plot_left
    plot_height = plot_bottom - plot_top

    def x_to_px(x_value: float) -> float:
        return plot_left + (x_value - x_min) / (x_max - x_min) * plot_width

    def y_to_px(y_value: float) -> float:
        return plot_bottom - (y_value - y_min) / (y_max - y_min) * plot_height

    image = Image.new("RGBA", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)
    font_tick = find_font(48)
    font_label = find_font(62)
    font_legend = find_font(34)
    axis_color = "#000000"
    label_color = "#000000"

    draw.line((plot_left, plot_bottom, plot_right, plot_bottom), fill=axis_color, width=6)
    draw.line((plot_left, plot_top, plot_left, plot_bottom), fill=axis_color, width=6)
    for tick in x_ticks:
        x = x_to_px(tick)
        draw.line((x, plot_bottom, x, plot_bottom + 12), fill=axis_color, width=5)
        draw_text_centered(draw, (x, plot_bottom + 58), format_tick(tick), font_tick, label_color)
    for tick in y_ticks:
        y = y_to_px(tick)
        draw.line((plot_left - 12, y, plot_left, y), fill=axis_color, width=5)
        label = format_tick(tick)
        bbox = draw.textbbox((0, 0), label, font=font_tick)
        draw.text((plot_left - 32 - (bbox[2] - bbox[0]), y - (bbox[3] - bbox[1]) / 2), label, font=font_tick, fill=label_color)

    draw_text_centered(draw, ((plot_left + plot_right) / 2, height - 68), "Current density / mA cm-2", font_label, label_color)
    draw_rotated_label(image, "Overpotential / mV", (76, (plot_top + plot_bottom) // 2), font_label, label_color)

    for target_j in target_current_densities:
        if x_min <= target_j <= x_max:
            x = x_to_px(target_j)
            for y0 in range(int(plot_top), int(plot_bottom), 28):
                draw.line((x, y0, x, min(y0 + 14, plot_bottom)), fill="#888888", width=3)

    for sample, color, rows in series:
        points_px = [(x_to_px(x), y_to_px(y)) for x, y in rows if x_min <= x <= x_max and y_min <= y <= y_max]
        if len(points_px) >= 2:
            draw.line(points_px, fill=color, width=8)

    target_by_sample: dict[str, list[dict[str, Any]]] = {}
    for row in target_rows:
        target_by_sample.setdefault(row["Sample"], []).append(row)
    for index, (sample, color, _rows) in enumerate(series):
        for row in target_by_sample.get(sample, []):
            if row["Status"] == "OK" and row["Eta_mV"] is not None:
                x, y = x_to_px(row["Target_j_mA_cm2"]), y_to_px(row["Eta_mV"])
                draw.ellipse((x - 9, y - 9, x + 9, y + 9), fill=color, outline="#000000", width=2)

    legend_x, legend_y = 340, 210
    for index, (sample, color, _rows) in enumerate(series):
        y = legend_y + index * 78
        labels = []
        for row in target_by_sample.get(sample, []):
            target_label = int(row["Target_j_mA_cm2"]) if float(row["Target_j_mA_cm2"]).is_integer() else row["Target_j_mA_cm2"]
            labels.append(f"eta{target_label}={row['Eta_mV']:.0f}" if row["Eta_mV"] is not None else f"eta{target_label}=NaN")
        draw.line((legend_x, y, legend_x + 120, y), fill=color, width=9)
        draw.text((legend_x + 145, y - 24), f"{sample}: " + ", ".join(labels), font=font_legend, fill=label_color)

    output = io.BytesIO()
    image.convert("RGB").save(output, format="PNG", optimize=True)
    return output.getvalue()


def add_overpotential_results_to_workbook(
    xlsx_path: Path,
    processed_rows: list[dict[str, Any]],
    target_rows: list[dict[str, Any]],
    summaries: list[dict[str, Any]],
    plot_png: bytes,
) -> None:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font

    workbook = load_workbook(xlsx_path)
    for sheet_name in ("Overpotential_Targets", "Overpotential_Data", "Summary", "Figure"):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    targets_sheet = workbook.create_sheet("Overpotential_Targets")
    target_headers = [
        "Sample",
        "Target_j_mA_cm2",
        "Current_A_equivalent",
        "E_RHE_iRcorr_V",
        "Eta_mV",
        "Eta_V",
        "pH",
        "Area_cm2",
        "Reference",
        "Ref_to_SHE_V",
        "iR_corrected",
        "Rs_ohm",
        "Status",
    ]
    targets_sheet.append(target_headers)
    for cell in targets_sheet[1]:
        cell.font = Font(bold=True)
    for row in target_rows:
        targets_sheet.append([row.get(header) for header in target_headers])
    targets_sheet.freeze_panes = "A2"

    data_sheet = workbook.create_sheet("Overpotential_Data")
    data_headers = [
        "Sample", "Pt", "Time_s", "E_ref_V", "I_A", "Current_sign_multiplier", "j_mA_cm2",
        "E_RHE_V", "E_RHE_iRcorr_V", "Eta_V", "Eta_mV", "OER_positive_branch", "Status"
    ]
    data_sheet.append(data_headers)
    for cell in data_sheet[1]:
        cell.font = Font(bold=True)
    for row in processed_rows:
        data_sheet.append([row.get(header) for header in data_headers])
    data_sheet.freeze_panes = "A2"

    summary_sheet = workbook.create_sheet("Summary")
    extra_headers: list[str] = []
    for summary in summaries:
        for key in summary:
            if key not in {"Sample", "Overpotential_note"} and key not in extra_headers:
                extra_headers.append(key)
    summary_headers = ["Sample"] + extra_headers + ["Overpotential_note"]
    summary_sheet.append(summary_headers)
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    for summary in summaries:
        summary_sheet.append([summary.get(header) for header in summary_headers])
    summary_sheet.freeze_panes = "A2"

    figure_sheet = workbook.create_sheet("Figure")
    figure_sheet["A1"] = "LSV overpotential target extraction"
    figure_sheet["A1"].font = Font(bold=True)
    image_path = xlsx_path.parent / "overpotential_plot_for_excel.png"
    image_path.write_bytes(plot_png)
    image = XLImage(str(image_path))
    image.width = 900
    image.height = 700
    figure_sheet.add_image(image, "A3")

    for sheet in (targets_sheet, data_sheet, summary_sheet):
        for column_cells in sheet.columns:
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(len(str(column_cells[0].value or "")) + 2, 12), 28)
    workbook.save(xlsx_path)
    image_path.unlink(missing_ok=True)
def make_response_zip(files: dict[str, bytes]) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for name, content in files.items():
            archive.writestr(name, content)
    return buffer.getvalue()


def parse_form(handler: BaseHTTPRequestHandler) -> tuple[list[dict[str, Any]], list[tuple[str, bytes]], dict[str, float]]:
    form = cgi.FieldStorage(
        fp=handler.rfile,
        headers=handler.headers,
        environ={
            "REQUEST_METHOD": "POST",
            "CONTENT_TYPE": handler.headers.get("Content-Type", ""),
        },
        keep_blank_values=True,
    )

    manifest_raw = form.getfirst("manifest", "[]")
    manifest = json.loads(manifest_raw)
    target_text = form.getfirst("target_current_densities", "100,200,300")
    target_current_densities = [float(part.strip()) for part in target_text.split(",") if part.strip()]
    if not target_current_densities:
        raise ValueError("At least one target current density is required.")

    reference_electrode = form.getfirst("reference_electrode", "Ag/AgCl")
    ref_to_she_text = form.getfirst("ref_to_she_v", "")
    if not ref_to_she_text:
        ref_to_she_text = "0.098" if reference_electrode == "Hg/HgO" else "0.197"

    options = {
        "tafel_min_j": float(form.getfirst("tafel_min_j", "10")),
        "tafel_max_j": float(form.getfirst("tafel_max_j", "100")),
        "reference_electrode": reference_electrode,
        "ref_to_she_v": float(ref_to_she_text),
        "target_current_densities": target_current_densities,
        "apply_ir_correction": form.getfirst("apply_ir_correction", "false").lower() in {"1", "true", "yes", "on"},
        "rs_ohm": float(form.getfirst("rs_ohm", "0") or 0),
        "ir_compensation_fraction": float(form.getfirst("ir_compensation_fraction", "1.0") or 1.0),
        "current_sign_mode": form.getfirst("current_sign_mode", "auto"),
    }
    uploaded = form["files"] if "files" in form else []
    if not isinstance(uploaded, list):
        uploaded = [uploaded]

    files: list[tuple[str, bytes]] = []
    for item in uploaded:
        if not getattr(item, "filename", ""):
            continue
        files.append((Path(item.filename).name, item.file.read()))

    return manifest, files, options


class Handler(BaseHTTPRequestHandler):
    server_version = "LSVTool/1.0"

    def log_message(self, format: str, *args: Any) -> None:
        timestamp = time.strftime("%H:%M:%S")
        print(f"[{timestamp}] {self.address_string()} {format % args}")

    def send_bytes(self, status: int, content_type: str, body: bytes, filename: str | None = None) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        if filename:
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        if self.path not in ("/", "/index.html"):
            self.send_bytes(404, "text/plain; charset=utf-8", "Not found".encode("utf-8"))
            return
        self.send_bytes(200, "text/html; charset=utf-8", HTML.encode("utf-8"))

    def do_POST(self) -> None:
        if self.path != "/api/export":
            self.send_bytes(404, "text/plain; charset=utf-8", "Not found".encode("utf-8"))
            return

        try:
            manifest, uploaded_files, options = parse_form(self)
            meta_by_name = {
                item["name"]: CurveMeta(
                    ph=float(item["pH"]),
                    area_cm2=float(item["area_cm2"]),
                    sheet_name=(item.get("sheet_name") or item["name"]),
                    reference_electrode=options["reference_electrode"],
                    ref_to_she_v=options["ref_to_she_v"],
                    apply_ir_correction=options["apply_ir_correction"],
                    rs_ohm=options["rs_ohm"],
                    ir_compensation_fraction=options["ir_compensation_fraction"],
                    current_sign_mode=options["current_sign_mode"],
                )
                for item in manifest
            }

            if not uploaded_files:
                raise ValueError("No files were received.")

            curves: list[tuple[Path, list[Any], CurveMeta]] = []
            WORKSPACE_TMP.mkdir(parents=True, exist_ok=True)
            run_dir = WORKSPACE_TMP / f"run_{int(time.time() * 1000)}"
            run_dir.mkdir()
            try:
                tmp = run_dir
                for filename, content in uploaded_files:
                    if filename not in meta_by_name:
                        raise ValueError(f"Missing parameters for {filename}")
                    source = tmp / filename
                    source.write_bytes(content)
                    points = extract_vfim(source)
                    curves.append((Path(filename), points, meta_by_name[filename]))

                full_path = tmp / "LSV_full.xlsx"
                workbook_resistance = options["rs_ohm"] * options["ir_compensation_fraction"] if options["apply_ir_correction"] else 0.0
                write_xlsx(full_path, curves, resistance_ohm=workbook_resistance, rhe_offset_v=options["ref_to_she_v"], ph_slope_v=0.0591)
                origin_xlsx, origin_csv = build_origin_tables(curves, resistance_ohm=workbook_resistance, rhe_offset_v=options["ref_to_she_v"], ph_slope_v=0.0591)
                plot_png = build_lsv_plot_png(curves, resistance_ohm=workbook_resistance, rhe_offset_v=options["ref_to_she_v"], ph_slope_v=0.0591)

                all_overpotential_rows: list[dict[str, Any]] = []
                all_target_rows: list[dict[str, Any]] = []
                overpotential_summaries: list[dict[str, Any]] = []
                processed_by_sample: dict[str, list[dict[str, Any]]] = {}
                used_overpotential_titles: set[str] = set()
                for source, points, meta in curves:
                    sample_name = sheet_title(meta.sheet_name or source.name, used_overpotential_titles)
                    processed_rows, target_rows, summary = extract_overpotentials_at_targets(
                        points,
                        sample_name,
                        meta,
                        options["target_current_densities"],
                    )
                    all_overpotential_rows.extend(processed_rows)
                    all_target_rows.extend(target_rows)
                    overpotential_summaries.append(summary)
                    processed_by_sample[sample_name] = processed_rows
                overpotential_png = build_overpotential_plot_png(
                    processed_by_sample,
                    all_target_rows,
                    options["target_current_densities"],
                )
                add_overpotential_results_to_workbook(
                    full_path,
                    all_overpotential_rows,
                    all_target_rows,
                    overpotential_summaries,
                    overpotential_png,
                )
                tafel_fits = compute_tafel_fits(
                    curves,
                    min_j=options["tafel_min_j"],
                    max_j=options["tafel_max_j"],
                    rhe_offset_v=options["ref_to_she_v"],
                    ph_slope_v=0.0591,
                )
                tafel_xlsx = build_tafel_workbook(tafel_fits)
                tafel_png = build_tafel_plot_png(tafel_fits)
                archive = make_response_zip(
                    {
                        "LSV_full.xlsx": full_path.read_bytes(),
                        "Origin_CE.xlsx": origin_xlsx,
                        "Origin_CE.csv": origin_csv,
                        "LSV_plot.png": plot_png,
                        "Overpotential_Targets_plot.png": overpotential_png,
                        "Tafel_fit.xlsx": tafel_xlsx,
                        "Tafel_plot.png": tafel_png,
                    }
                )
            finally:
                for child in sorted(run_dir.rglob("*"), reverse=True):
                    if child.is_file():
                        child.unlink(missing_ok=True)
                    elif child.is_dir():
                        child.rmdir()
                run_dir.rmdir()

            self.send_bytes(200, "application/zip", archive, filename="LSV_export.zip")
        except Exception as exc:  # noqa: BLE001 - user-facing local tool error.
            message = str(exc).encode("utf-8")
            self.send_bytes(400, "text/plain; charset=utf-8", message)


def main() -> int:
    parser = argparse.ArgumentParser(description="Start the local LSV conversion web app.")
    parser.add_argument("--host", default=HOST, help="Host to bind. Default: 127.0.0.1")
    parser.add_argument("--port", type=int, default=PORT, help="Port to bind. Default: 8765")
    parser.add_argument("--no-open", action="store_true", help="Do not open the browser automatically.")
    args = parser.parse_args()

    server = ThreadingHTTPServer((args.host, args.port), Handler)
    url = f"http://{args.host}:{args.port}"
    print(f"LSV tool started: {url}")
    print("Press Ctrl+C to stop.")
    if not args.no_open:
        webbrowser.open(url)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped.")
    finally:
        server.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
