#!/usr/bin/env python3
"""
Extract Vf/Im data from Gamry-style LSV .DTA files.

Examples:
  python scripts/extract_lsv_vfim.py "C:\\data\\one.DTA.###"
  python scripts/extract_lsv_vfim.py "C:\\data" --recursive
  python scripts/extract_lsv_vfim.py "C:\\data" --recursive --combined all_lsv.csv
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


TABLE_RE = re.compile(r"^CURVE\s+TABLE\s+(\d+)", re.IGNORECASE)


@dataclass
class LsvPoint:
    source_file: str
    point: int
    time_s: float
    vf_v: float
    im_a: float


@dataclass
class CurveMeta:
    ph: float
    area_cm2: float
    sheet_name: str | None = None
    reference_electrode: str = "Ag/AgCl"
    ref_to_she_v: float = 0.197
    apply_ir_correction: bool = False
    rs_ohm: float = 0.0
    ir_compensation_fraction: float = 1.0
    current_sign_mode: str = "auto"


def read_text(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gbk", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("latin-1", errors="replace")


def split_fields(line: str) -> list[str]:
    return [field.strip() for field in line.strip().split("\t")]


def parse_float(value: str) -> float:
    return float(value.replace(",", ""))


def find_curve_table(lines: list[str]) -> tuple[int, int]:
    for index, line in enumerate(lines):
        match = TABLE_RE.match(line.strip())
        if match:
            return index, int(match.group(1))
    raise ValueError("No CURVE TABLE section found")


def extract_vfim(path: Path) -> list[LsvPoint]:
    text = read_text(path)
    lines = text.splitlines()
    table_index, expected_rows = find_curve_table(lines)

    if table_index + 3 >= len(lines):
        raise ValueError("CURVE TABLE section is incomplete")

    headers = split_fields(lines[table_index + 1])
    try:
        point_col = headers.index("Pt")
        time_col = headers.index("T")
        vf_col = headers.index("Vf")
        im_col = headers.index("Im")
    except ValueError as exc:
        raise ValueError(f"Required columns not found in CURVE header: {headers}") from exc

    points: list[LsvPoint] = []
    for line in lines[table_index + 3 :]:
        fields = split_fields(line)
        if len(fields) <= max(point_col, time_col, vf_col, im_col):
            break
        if not fields[point_col].isdigit():
            break

        points.append(
            LsvPoint(
                source_file=str(path),
                point=int(fields[point_col]),
                time_s=parse_float(fields[time_col]),
                vf_v=parse_float(fields[vf_col]),
                im_a=parse_float(fields[im_col]),
            )
        )

        if len(points) >= expected_rows:
            break

    if not points:
        raise ValueError("CURVE TABLE contains no parseable data rows")

    return points


def discover_inputs(paths: Iterable[Path], recursive: bool, pattern: str) -> list[Path]:
    files: list[Path] = []
    for path in paths:
        if path.is_file():
            files.append(path)
        elif path.is_dir():
            iterator = path.rglob(pattern) if recursive else path.glob(pattern)
            files.extend(p for p in iterator if p.is_file())
        else:
            raise FileNotFoundError(path)

    return sorted(dict.fromkeys(p.resolve() for p in files))


def safe_output_name(path: Path) -> str:
    name = path.name
    for suffix in path.suffixes:
        name = name.removesuffix(suffix)
    if not name:
        name = path.stem or "lsv"
    return f"{name}_Vf_Im.csv"


def write_points_csv(path: Path, points: Iterable[LsvPoint]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["source_file", "Pt", "T_s", "Vf_V", "Im_A"])
        for point in points:
            writer.writerow(
                [
                    point.source_file,
                    point.point,
                    f"{point.time_s:.12g}",
                    f"{point.vf_v:.12g}",
                    f"{point.im_a:.12g}",
                ]
            )


def read_metadata_csv(path: Path) -> dict[str, CurveMeta]:
    metadata: dict[str, CurveMeta] = {}
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        normalized = {field.lower(): field for field in fieldnames}
        required = {"file", "ph", "area_cm2"}
        missing = required - set(normalized)
        if missing:
            raise ValueError(f"Metadata CSV is missing columns: {', '.join(sorted(missing))}")

        for row in reader:
            file_value = (row.get(normalized["file"]) or "").strip()
            if not file_value:
                continue
            meta = CurveMeta(
                ph=float((row.get(normalized["ph"]) or "").strip()),
                area_cm2=float((row.get(normalized["area_cm2"]) or "").strip()),
                sheet_name=(row.get(normalized.get("sheet_name", "")) or "").strip() or None,
            )
            metadata[file_value] = meta
            metadata[Path(file_value).name] = meta
            try:
                metadata[str(Path(file_value).resolve())] = meta
            except OSError:
                pass
    return metadata


def write_metadata_template(path: Path, files: Iterable[Path]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(["file", "pH", "area_cm2", "sheet_name"])
        for source in files:
            writer.writerow([str(source), "", "", source.stem[:31]])


def metadata_for_file(
    path: Path,
    metadata: dict[str, CurveMeta],
    default_ph: float | None,
    default_area: float | None,
    prompt: bool,
) -> CurveMeta:
    keys = (str(path), str(path.resolve()), path.name)
    for key in keys:
        if key in metadata:
            return metadata[key]

    if default_ph is not None and default_area is not None:
        return CurveMeta(ph=default_ph, area_cm2=default_area)

    if not prompt:
        raise ValueError(
            f"No pH/area metadata for {path}. Use --ph and --area, or provide --metadata."
        )

    print(f"\nMetadata for {path.name}")
    ph = default_ph
    area = default_area
    while ph is None:
        value = input("  pH: ").strip()
        try:
            ph = float(value)
        except ValueError:
            print("  Please enter a number, for example 8.4")
    while area is None:
        value = input("  area_cm2: ").strip()
        try:
            area = float(value)
        except ValueError:
            print("  Please enter a number, for example 0.25")

    return CurveMeta(ph=ph, area_cm2=area)


def sheet_title(base: str, used: set[str]) -> str:
    invalid = r"[]:*?/\\"
    cleaned = "".join("_" if char in invalid else char for char in base).strip()
    cleaned = cleaned[:31] or "LSV"
    title = cleaned
    counter = 2
    while title in used:
        suffix = f"_{counter}"
        title = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(title)
    return title


def write_xlsx(
    path: Path,
    curves: list[tuple[Path, list[LsvPoint], CurveMeta]],
    resistance_ohm: float,
    rhe_offset_v: float,
    ph_slope_v: float,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError(
            "Writing .xlsx requires openpyxl. Install openpyxl or run with the Codex bundled Python."
        ) from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    inputs = workbook.active
    inputs.title = "_inputs"
    inputs.append(
        [
            "sheet_name",
            "source_file",
            "pH",
            "area_cm2",
            "rhe_shift_V",
            "resistance_ohm",
            "rows",
        ]
    )
    for cell in inputs[1]:
        cell.font = Font(bold=True)

    used_titles = {inputs.title}
    for index, (source, points, meta) in enumerate(curves, start=2):
        preferred = meta.sheet_name or source.name
        sheet_name = sheet_title(preferred, used_titles)
        rhe_shift = rhe_offset_v + ph_slope_v * meta.ph

        inputs.append(
            [
                sheet_name,
                str(source),
                meta.ph,
                meta.area_cm2,
                rhe_shift,
                resistance_ohm,
                len(points),
            ]
        )

        sheet = workbook.create_sheet(sheet_name)
        sheet.append(
            [
                "V vs. Ref.",
                "A",
                "v",
                "V-IR versus RHE V",
                "current density mA/cm2",
            ]
        )
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row_number, point in enumerate(points, start=2):
            sheet.cell(row=row_number, column=1, value=point.vf_v)
            sheet.cell(row=row_number, column=2, value=point.im_a)
            sheet.cell(row=row_number, column=3, value=f"=SUM(A{row_number},_inputs!E{index})")
            sheet.cell(
                row=row_number,
                column=4,
                value=f"=SUM(C{row_number},-B{row_number}*_inputs!F{index})",
            )
            sheet.cell(
                row=row_number,
                column=5,
                value=f"=SUM(B{row_number}*1000/_inputs!D{index})",
            )

        sheet.freeze_panes = "A2"
        widths = {"A": 14, "B": 14, "C": 14, "D": 20, "E": 22}
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

    for column, width in {"A": 18, "B": 70, "C": 10, "D": 12, "E": 13, "F": 15, "G": 10}.items():
        inputs.column_dimensions[column].width = width

    workbook.save(path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract Vf and Im columns from LSV CURVE TABLE data."
    )
    parser.add_argument("inputs", nargs="+", type=Path, help="Input file(s) or folder(s).")
    parser.add_argument(
        "-o",
        "--out-dir",
        type=Path,
        default=Path("lsv_vfim_output"),
        help="Folder for per-file CSV outputs. Default: lsv_vfim_output",
    )
    parser.add_argument(
        "--combined",
        type=Path,
        help="Optional single CSV containing all extracted files.",
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        help="Optional Excel workbook in the same format as the LSV template.",
    )
    parser.add_argument(
        "--metadata",
        type=Path,
        help="CSV with columns: file,pH,area_cm2. Optional column: sheet_name.",
    )
    parser.add_argument(
        "--write-metadata-template",
        type=Path,
        help="Write a metadata CSV template for discovered files, then exit.",
    )
    parser.add_argument("--ph", type=float, help="pH used for all curves when --metadata is not used.")
    parser.add_argument(
        "--area",
        type=float,
        help="Electrode area in cm^2 used for all curves when --metadata is not used.",
    )
    parser.add_argument(
        "--no-prompt",
        action="store_true",
        help="Do not prompt for missing pH/area values.",
    )
    parser.add_argument(
        "--resistance",
        type=float,
        default=3.8504,
        help="Resistance for IR correction in ohm. Default: 3.8504",
    )
    parser.add_argument(
        "--rhe-offset",
        type=float,
        default=0.197,
        help="Reference offset in V for RHE conversion. Default: 0.197",
    )
    parser.add_argument(
        "--ph-slope",
        type=float,
        default=0.0591,
        help="pH slope in V/pH for RHE conversion. Default: 0.0591",
    )
    parser.add_argument(
        "-r",
        "--recursive",
        action="store_true",
        help="Search input folders recursively.",
    )
    parser.add_argument(
        "--pattern",
        default="*.DTA*",
        help="File pattern used when an input is a folder. Default: *.DTA*",
    )
    parser.add_argument(
        "--no-per-file",
        action="store_true",
        help="Do not write one CSV per source file. Useful with --combined.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    files = discover_inputs(args.inputs, args.recursive, args.pattern)
    if not files:
        parser.error("No input files found.")

    if args.write_metadata_template:
        write_metadata_template(args.write_metadata_template, files)
        print(f"METADATA -> {args.write_metadata_template} ({len(files)} files)")
        return 0

    combined_points: list[LsvPoint] = []
    extracted_curves: list[tuple[Path, list[LsvPoint], CurveMeta]] = []
    failed: list[tuple[Path, str]] = []
    metadata = read_metadata_csv(args.metadata) if args.metadata else {}

    for source in files:
        try:
            points = extract_vfim(source)
            if args.xlsx:
                meta = metadata_for_file(
                    source,
                    metadata,
                    args.ph,
                    args.area,
                    prompt=not args.no_prompt,
                )
                extracted_curves.append((source, points, meta))
        except Exception as exc:  # noqa: BLE001 - keep batch mode moving.
            failed.append((source, str(exc)))
            continue

        combined_points.extend(points)
        if not args.no_per_file:
            output_path = args.out_dir / safe_output_name(source)
            write_points_csv(output_path, points)
            print(f"OK  {source} -> {output_path} ({len(points)} rows)")
        else:
            print(f"OK  {source} ({len(points)} rows)")

    if args.xlsx:
        if not extracted_curves:
            print("No data extracted; Excel workbook was not written.")
        else:
            write_xlsx(
                args.xlsx,
                extracted_curves,
                resistance_ohm=args.resistance,
                rhe_offset_v=args.rhe_offset,
                ph_slope_v=args.ph_slope,
            )
            print(f"XLSX -> {args.xlsx} ({len(extracted_curves)} sheets)")

    if args.combined:
        if not combined_points:
            print("No data extracted; combined CSV was not written.")
        else:
            write_points_csv(args.combined, combined_points)
            print(f"ALL -> {args.combined} ({len(combined_points)} rows)")

    if failed:
        print("\nFailed files:")
        for source, reason in failed:
            print(f"FAIL {source}: {reason}")
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
